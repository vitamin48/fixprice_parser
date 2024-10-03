"""
Скрипт считывает файл JSON с товарами Офисмаг и записывает данные в Excel.
"""
import json
import re
import pandas as pd

from openpyxl.utils import get_column_letter

FILE_NAME_JSON = 'merge_dictionaries/result_merge/data.json'


def read_json():
    with open(FILE_NAME_JSON, 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        return data


def read_bad_brand():
    """Считывает и возвращает список нежелательных брендов"""
    with open('in/bad_brand.txt', 'r', encoding='utf-8') as file:
        brands = [line.strip().lower() for line in file if line.strip()]
    return set(brands)


def contains_unwanted_brand(product_name, bad_brands):
    """Функция для проверки наличия нежелательных брендов в имени товара"""
    for brand in bad_brands:
        if brand in product_name.strip().lower():
            return True
    return False


def extract_numbers(input_string):
    return int(''.join([char for char in str(input_string) if char.isdigit()]))


def transform_price(x):
    result = x * 5 if x < 200 else (
        x * 4.5 if 200 <= x < 500 else (
            x * 4 if 500 <= x < 1000 else (
                x * 3.5 if 1000 <= x < 5000 else (
                    x * 3 if 5000 <= x < 10000 else (
                        x * 2.5 if 10000 <= x < 20000 else (x * 2))))))
    # Убеждаемся, что значение после преобразований не меньше 490
    result = max(result, 490)
    # Округление до целого числа
    return round(result)


def create_rows_for_df_by_dict(data_dict):
    bad_brands = read_bad_brand()
    rows_main = []
    rows_stock = []
    # Проходим по каждому ключу в словаре
    for key, value in data_dict.items():
        """Обработка бренда"""
        brand = value.get('brand', 'NO_KEY')
        if brand.strip().lower() in bad_brands:
            continue
        name = value.get('name')
        """Обработка бренда в имени"""
        if contains_unwanted_brand(product_name=name, bad_brands=bad_brands):
            continue
        """Обработка характеристик"""
        # characteristics = value.get("characteristics", {})
        "Страна"
        country = value.get('country', 'NO_KEY')
        "Высота х Длина х Ширина"
        height = value.get('packing_height', value.get('height', 'NO_KEY'))
        length = value.get('package_length', value.get('length', 'NO_KEY'))
        width = value.get('packing_width', value.get('width', 'NO_KEY'))
        "Цена"
        price = value.get('price', 'NO_KEY')
        modified_price = round(float(price))
        "Остатки"
        stock = value.get('stock')
        "Вес"
        weight = value.get('weight')
        "Описание"
        description = value.get("description", "NO_KEY")
        "Изображения"
        img_urls = value.get("url_img", [])
        if isinstance(img_urls, str):
            img_urls = [img_urls]
        if len(img_urls) > 0:  # Извлечение первой ссылки и всех остальных
            img_url1 = img_urls[0]
            img_url2 = img_urls[1:]  # Все остальные ссылки
            # Преобразуем список ссылок в строку, разделенную запятой, или оставляем как есть.
            img_url2 = ", ".join(img_url2) if len(img_url2) > 0 else "-"
        else:
            img_url1 = "-"
            img_url2 = "-"
        "Формируем итоговую главную строку"
        row = {
            "ArtNumber": key,
            "Название": name,
            "Цена FIX": modified_price,
            "Описание": description,
            "Ссылка на главное фото товара": img_url1,
            "Ссылки на другие фото товара": img_url2,
            # "art_url": value.get("art_url", ""),
            "Бренд": brand,
            "Страна": country,
            "Ширина, мм": width,
            "Высота, мм": height,
            "Длина, мм": length,
            "Вес": weight
        }
        "Формируем строку с остатками"
        row_stock = {
            "ArtNumber": f'p_{key}',
            "Название": name,
            "Остатки": stock
        }
        rows_main.append(row)
        rows_stock.append(row_stock)
    return rows_main, rows_stock


def create_df_by_rows(rows_main, rows_stock):
    # Создание DataFrame из списка словарей
    df_main = pd.DataFrame(rows_main)
    df_stock = pd.DataFrame(rows_stock)
    # Добавляем артикул
    df_main["Артикул"] = df_main["ArtNumber"].apply(lambda art: f'p_{art}')
    # Добавляем столбец Цена для OZON
    df_main['Цена для OZON'] = df_main['Цена FIX'].apply(transform_price)
    # Добавляем столбец Цена до скидки
    df_main['Цена до скидки'] = df_main['Цена для OZON'].apply(lambda x: int(round(x * 1.3)))
    # Добавляем столбец НДС Не облагается
    df_main["НДС"] = "Не облагается"
    # Задаем порядок столбцов
    desired_order = ['Артикул', 'Название', 'Цена для OZON', 'Цена до скидки', 'НДС', 'Цена FIX', 'Вес',
                     'Ширина, мм', 'Высота, мм', 'Длина, мм', 'Ссылка на главное фото товара',
                     'Ссылки на другие фото товара', 'Бренд', 'ArtNumber', 'Описание', 'Страна']
    result_main_df = df_main[desired_order]
    return result_main_df, df_stock


def create_xls(df_main, df_stock):
    file_name = f'FP_2559.xlsx'
    # Сохранение DataFrame в Excel с использованием Styler
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df_main.to_excel(writer, sheet_name='Данные', index=False, na_rep='NaN')
        df_stock.to_excel(writer, sheet_name='Остатки', index=False, na_rep='NaN')
        "Работа с данными"
        # Установка ширины столбцов
        worksheet_ozon = writer.sheets['Данные']
        for column in df_main:
            column_width = max(df_main[column].astype(str).map(len).max(), len(column)) + 2
            col_letter = get_column_letter(df_main.columns.get_loc(column) + 1)
            worksheet_ozon.column_dimensions[col_letter].width = column_width
        # Закрепите первую строку
        worksheet_ozon.freeze_panes = 'A2'
        # Корректировка ширины столбцов
        worksheet_ozon.column_dimensions[get_column_letter(df_main.columns.get_loc('Название') + 1)].width = 30
        worksheet_ozon.column_dimensions[get_column_letter(df_main.columns.get_loc('Описание') + 1)].width = 30
        worksheet_ozon.column_dimensions[
            get_column_letter(df_main.columns.get_loc('Ссылка на главное фото товара') + 1)].width = 30
        worksheet_ozon.column_dimensions[
            get_column_letter(df_main.columns.get_loc('Ссылки на другие фото товара') + 1)].width = 30
        "Работа с остатками"
        worksheet_srock = writer.sheets['Остатки']
        for column in df_stock:
            column_width = max(df_stock[column].astype(str).map(len).max(), len(column)) + 2
            col_letter = get_column_letter(df_stock.columns.get_loc(column) + 1)
            worksheet_srock.column_dimensions[col_letter].width = column_width
        # Корректировка ширины столбцов
        worksheet_srock.column_dimensions[get_column_letter(df_main.columns.get_loc('Название') + 1)].width = 30


if __name__ == '__main__':
    data_json = read_json()
    rows_main, rows_stock = create_rows_for_df_by_dict(data_dict=data_json)
    df_main, df_stock = create_df_by_rows(rows_main, rows_stock)
    create_xls(df_main, df_stock)
